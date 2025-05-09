{
 "cells": [
  {
   "cell_type": "markdown",
   "metadata": {},
   "source": [
    "# Overview"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": null,
   "metadata": {},
   "outputs": [],
   "source": [
    "# Import libraries\n",
    "# yfinance offers a reliable, threaded, and Pythonic way to download historical market data from Yahoo! finance\n",
    "# Please check out its official doc for details: https://pypi.org/project/yfinance/\n",
    "import yfinance as yf\n",
    "import pandas as pd\n",
    "\n",
    "# Load historical data in the past 10 years\n",
    "sp500 = yf.Ticker(\"^GSPC\")\n",
    "end_date = pd.Timestamp.today()\n",
    "start_date = end_date - pd.Timedelta(days=10*365)\n",
    "sp500_history=sp500.history(start=start_date, end=end_date)\n",
    "\n",
    "# Remove unnecessary columns\n",
    "sp500_history = sp500_history.drop(columns=['Dividends', 'Stock Splits'])\n",
    "\n",
    "# Create a new column as Close 200 days moving average\n",
    "sp500_history['Close_200ma'] = sp500_history['Close'].rolling(200).mean()\n",
    "\n",
    "# Create a summary statistics table\n",
    "sp500_history_summary = sp500_history.describe()"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": null,
   "metadata": {},
   "outputs": [
    {
     "data": {
      "image/png": "iVBORw0KGgoAAAANSUhEUgAAAgsAAADeCAYAAACzM0N7AAAAOXRFWHRTb2Z0d2FyZQBNYXRwbG90bGliIHZlcnNpb24zLjMuMywgaHR0cHM6Ly9tYXRwbG90bGliLm9yZy/Il7ecAAAACXBIWXMAAAsTAAALEwEAmpwYAABE2ElEQVR4nO3dd3hU1dbA4d9KT0gIAUIICRBC7y1SBBTEgqAiKooFO1y7XNsVy6de5VqvBRW92AAboqKAiCggikrH0EGChBogtCQQ0vf3xzlJZpJJL5OE9T7PPJmzzz5n1gQyWdlnn7XFGINSSimlVFE83B2AUkoppWo2TRaUUkopVSxNFpRSSilVLE0WlFJKKVUsTRaUUkopVSxNFpRSSilVrFInCyLiKSJ/ish39vbTIrJfRGLtx3CHvhNFJE5EtovIRQ7tvUVko71vsohISa87bNgwA+hDH/rQhz7qxkPVQmUZWbgf2Fqg7TVjTA/78T2AiHQCxgCdgWHAFBHxtPu/A4wH2tqPYSW96JEjR8oQolJKKaUqW6mSBRGJBEYA75ei+0hgpjEm3RizC4gD+ohIOFDfGLPcWJWgZgCXly9spZRSSlWX0o4svA48AuQUaL9HRDaIyIciEmK3RQB7Hfrss9si7OcF2wsRkfEiskZE1iQmJpYyRKWUUkpVhRKTBRG5BDhsjFlbYNc7QGugB5AA/Df3EBenMcW0F240ZqoxJsYYExMaGlpSiEoppZSqQl6l6DMAuMyewOgH1BeRT4wxN+R2EJH3gO/szX1Ac4fjI4EDdnuki3allFJK1WAljiwYYyYaYyKNMVFYExeXGGNusOcg5BoFbLKfzwXGiIiviLTCmsi4yhiTAKSISD/7LogbgTmV+WaUUkpVjS0HkklOy3R3GMpNKlJn4SX7NsgNwBDgnwDGmM3ALGAL8ANwtzEm2z7mTqxJknHATmBBBV5fKaVUNdhzNJXL3/6dl3/Y7u5QlJuU5jJEHmPMUmCp/XxsMf0mAZNctK8BupQpQqWUUm5z/FQGo//3BxnZOVzYOczd4Sg30QqOSimlijT6f8s5lJwOQFh9PzdHo9xFkwWllFIu7TmaStzhk3nbTYJ83RiNcidNFpRSSrm06+gpp+1A3zJduVZ1iCYLSimlXEo67Xz3g5en/so4U+m/vFJKKZf+iDuCt2eJ6/2pM4AmC0oppQrZfCCJmav3ckm3Zu4ORdUAmiwopZQiKzuH3fYchYSk04yY/BsA0Y3rceuAVnSNCHZneMrNdLaKUkopHpi1nrnrD/DHo+dx9gtL8tqbNfDn3qFt3RiZqgl0ZEEppc5w2w4mM3e9tVTPrdNWO+1r1sDfHSGpGkaTBaWUOoOt2nWMYa8vy9vedjDFaX+EJgsKTRaUUuqM9uWavYXa+rZqSMfw+gCEBWshJqVzFpRS6oz25dp9hdq6RARz5+DW/HUoBV8vTzdEpWoaTRaUUuoMlZmd47K9fVgQjQN9aRyoowrKopchlFLqDHU8NQOA7pHB7Hp+eF774A6h7gpJ1VCaLCil1Blq3e7jANw6sBUi+ZUamwTp6pLKWamTBRHxFJE/ReQ7e7uhiPwkIjvsryEOfSeKSJyIbBeRixzae4vIRnvfZHH836mUUqpardtzAoDOzerntYUEeLspGlWTlWVk4X5gq8P2o8BiY0xbYLG9jYh0AsYAnYFhwBQRyZ0h8w4wHmhrP4ZVKHqllFLldjI9i5AAb9o0CQLgs3F9WXD/OW6OStVEpUoWRCQSGAG879A8EphuP58OXO7QPtMYk26M2QXEAX1EJByob4xZbowxwAyHY5RSSlWzvcdSad4wIG/77NaNaRqslyBUYaUdWXgdeARwnDobZoxJALC/NrHbIwDHG3f32W0R9vOC7Uoppdwg+XQmIQE+7g5D1QIlJgsicglw2BiztpTndDUPwRTT7uo1x4vIGhFZk5iYWMqXVUopVRrJaZm0mjif9fuS8u6IUKo4pRlZGABcJiLxwEzgPBH5BDhkX1rA/nrY7r8PaO5wfCRwwG6PdNFeiDFmqjEmxhgTExqqt/AopVRlmjAzFmP/qbZhX5J7g1G1QonJgjFmojEm0hgThTVxcYkx5gZgLnCT3e0mYI79fC4wRkR8RaQV1kTGVfalihQR6WffBXGjwzFKKaWqUFJqJrdOW81rP/3lVIzpo1vOcmNUqraoSAXHF4BZInIbsAcYDWCM2Swis4AtQBZwtzEm2z7mTmAa4A8ssB9KKaWq2HPzt7Bk22GWbDvs1N6haVDJBx+Ph5VT4aJJoHe8n5HKlCwYY5YCS+3nR4GhRfSbBExy0b4G6FLWIJVSSlXMyl3HXLYH+JTi18DSF2DDLBg4AQKblNhd1T1awVEppc4AWS7WgejQNIj6fiUkC0d3wvqZ0PsmTRTOYJosKKVUHZaYkk7Uo/M5kJTG7QNb5bVfE9OcHyacQ4mFdFMSIPpcGPhAFUeqajJNFpRSqo46fiqDB2bF5m03qZ+/iuQPmw+WfILsLIgaCDfOgQbNS+6v6ixNFpRSqo7619cbWLbjSN52kF/+ug9h9UtYfjrjFLw/FE7sqarwVC2iyYJSStVRyWmZTttBDvMTJg7vWPSBxsB3D0BCLByNq6LoVG2iyYJSStVRzRr4O20H+XkzsE1jADyLm6uwdR5smAkD/wmtz6vKEFUtocmCUkrVQZsPJDF73X6ntkb1fPjPqK5c2r0ZfaMbuj7wxB6Ydx806QyDH6uGSFVtUJGiTEoppWqgw8lpjJj8W6H2JkG+NKnvx5vX9nR9oDEw917ISIWRb4GXLjKlLJosKKVUHRJ3+CTnv/qLU1u3yGA27EuiYb0SfvmLwPlPQ9I+iOhVdUGqWkeTBaWUqkMKJgrndwzjpau6sevISbw8i7nynJ4CmWnQrKf1UMqBJgtKKVVHfXv3AHo0bwBAw3pFzFHI9ctLsHcl3LpQ139QhegER6WUqiOOnkx32s5NFEp0Yi+smAJB4ZooKJc0WVBKqTpgy4Fkej+3KG/7g5tiSn/wkmfB5MAFz1RBZKou0MsQSilVBziWdf7yjv6cFVXCZYdchzbDxq8g5jYIiaqS2FTtpyMLSilVS5xIzeB/v+zEGFNo395jqXnPS50oAPz2Onj6wDkPV0KEqq7SkQWllKolHv9mE/M3JtCjeQP6Rjdy2ncqI7t8J+19M0QNgKCwigeo6qwSRxZExE9EVonIehHZLCLP2O1Pi8h+EYm1H8MdjpkoInEisl1ELnJo7y0iG+19k6XEtVGVUkrlSkyxJjDmjiscPZnOnqOppGXmJwoPXNCudCfLyoDsTCtR6H1z5Qaq6pzSXIZIB84zxnQHegDDRKSfve81Y0wP+/E9gIh0AsYAnYFhwBQR8bT7vwOMB9raj2GV9k6UUqqOS8+ykgJvT+vvrCGvLOWcl39m15FTeX3uG9q2dCdb+h94ZwCkJVV6nKruKTFZMJaT9qa3/Sh8wSzfSGCmMSbdGLMLiAP6iEg4UN8Ys9xYF9xmAJdXKHqllDqDrN/n/Is9OS0LgGU7EgGrrkKpHNxozVWI6A1+wZUZoqqjSjXBUUQ8RSQWOAz8ZIxZae+6R0Q2iMiHIhJit0UAex0O32e3RdjPC7a7er3xIrJGRNYkJiaW/t0opdQZIDPb+e+1XUesyY3RofVKPtgY+P5hK0k4/+kqiE7VRaVKFowx2caYHkAk1ihBF6xLCq2xLk0kAP+1u7uah2CKaXf1elONMTHGmJjQ0NDShKiUUnVaTk7+x2Vmdo7Tvs9X7QHAp7hyzrk2fQ17lsO5j+ikRlVqZbp10hhzAlgKDDPGHLKTiBzgPaCP3W0f0NzhsEjggN0e6aJdKaVUCY6nZuQ9z8jKcXn7pK9XCR/p2Znw45PQqA30vbOyQ1R1WGnuhggVkQb2c3/gfGCbPQch1yhgk/18LjBGRHxFpBXWRMZVxpgEIEVE+tl3QdwIzKm8t6KUUnWX4yTG9Kwc0jJzCvUp8QYzDy8YPQ2ufB88tMyOKr3S1FkIB6bbdzR4ALOMMd+JyMci0gPrUkI88A8AY8xmEZkFbAGygLuNMbn39dwJTAP8gQX2Qyml6rTsHENmdg5+3p4ldy7Csh1H8p7f9ek6ruodWUxvFzJPg6cvtOhb7hjUmUtcDWXVJDExMWbNmjXuDkMppcrlyMl0Yuw1G+JfGFGuc+w5mso5L//scp+Xh5Blz2co9vzf/RN2LoG7VoK3X7niqCRaX6cW0nEopZSqQr85jAg4Fk8qizs/XVvkvk7N6pd8ghN7YN0MiBro7kRB1VKaLCilVBXYf+I0i7ceYsIXsXltJ9OzynWuzQeSAfjxn+cU2vfosA4ln2DJJBAPGPDPcr2+Uro2hFJKVYGJszfy61/OdWJOpmXRONC33OdsHhLgtB3duB6BfiV8jB/eBhtmwlnjoHGbcr+2OrPpyIJSSlUBx0QhtzxzeUYWjp601oN4YkRH/H08neYlPHRRe46eyijqUMuip6y7IAY9WObXViqXJgtKKVXJNhYoyxzRwB+AU2VMFhJT0hnw4hIAukYULsvs7+NJZlbhWyidDHoIrp4B9cOL76dUMfQyhFJKVaKs7Bwufes3p7YgP2+g7CMLryzcnldPoaPDRMaWjQLYfTQVLw+hW0ur0v6gto1dn6T5WWV6TaVc0ZEFpZSqRLfPKHyrd9Ng6w6EkpKFwylpPL9gK1l2OecftxzM2xfkm/+3XZdm1ihDgI8njQJ9iX9hBB/fVqB+wqmj8PEo2PhVud6HUo40WVBKqUqy5UAyS7dbcxUiQ/yZNKoLAKPtAkqn0ou/dfLxbzbxv1/+ZlX8MQCOp2bm7XOszvjkJZ2YeHEHejYPKXSOPMvftOoqBOnlB1VxehlCKaUqyZzY/QA8Mqw9dw1ugzGGS7o1yxspyMopfn5BQtJpwFr7wbEmwxtjejj1axrsxz/ObV30iRK3w/K3ofVQaHl2Od6JUs50ZEEppSpJbj2EcYOiAWs0INjfG297gaeMEiYj5uYSKWlZdHjyBwCGdW7KyB4RpQ/iZCJ8MRZ86sHIt6Ck9SKUKgUdWVBKqQradjCZCTNj2XYwhbH9WuJdYKno3KWjM7OLL6+fY5ffz3Eow5+RXcLdDgX98C+rYuMNX0P9ZmU7Vqki6MiCUkpVQEZWDjd/uJptB1MACKtfuOhSbvLw4g/bOJ1R9LyF3JUl758Zm9f24pXdin7xzDTY8CV8dRvssydW9r0TbpwDUQPK+E6UKpqOLCilVAW8/9vfHExOy9seHdO8UB9Pj/xLAZ+s2M24c6Jdniu9wGWKt6/rRWhQgeQjJwfifoLtC2DzN5B2AgLDoFEbiIzRWyVVldBkQSmlyun4qQxe+mG7U1tY/eIXairLZYVCiQLAjh/h82vAJxDaXQS9boKoQeChA8Wq6miyoJRS5fTU3M0AXNylKaN6RhBSz6fEY46VVJ7ZQVDuug9pSbBtPvS4DqLPhas+go6Xgqd3ueJWqqxKTBZExA/4FfC1+39ljHlKRBoCXwBRQDxwtTHmuH3MROA2IBu4zxiz0G7vDUwD/IHvgfuNMcXP+FFKqRpq7voDAEy5vpdTHYRC0pIZ5/kd9SSNJPOQyy5JDjUVcgX5esCGWbDwcUg9Cq3Pg6Cm0OWKSolfqdIqzbhVOnCeMaY70AMYJiL9gEeBxcaYtsBiexsR6QSMAToDw4ApIuJpn+sdYDzQ1n4Mq7y3opRSlSM9K5snvt3IT1sOFdln0/789R+KTRQO/AlT+vO492ec67GB9Mxs626FZa9C8oG8bhO++DPveXeJ417P2TSbeSHMHgfBETBusZUoKOUGJSYLxnLS3vS2HwYYCUy326cDl9vPRwIzjTHpxphdQBzQR0TCgfrGmOX2aMIMh2OUUqpG2LgvifZP/MAnK/YwzqF089rdx4h6dD6frtxtFVt601r/4bt7B7o+UXYW/PYavH8BmGzGZDzBqIx/k55lIG4xLH4GXu0Ir3WBr27j8rRvaSlWeeezPbbwoPdXeHh4wqj/we1LoFnPKn/vShWlVHMW7JGBtUAb4G1jzEoRCTPGJAAYYxJEpIndPQJY4XD4Prst035esN3V643HGoGgRYsWpX83SilVAaczsgstApWelY2vlyeT5m8FrJLMf8QdzdvfxcVqkADEfgKLnoZ2F8Olb7BikpV4tAsLhN43Q3g32LMS9q6A+N8YefIgcR5X8Wb2FSS0u57Z7R/gin7tq+JtKlVmpUoWjDHZQA8RaQB8IyJdiunuajzOFNPu6vWmAlMBYmJidE6DUqparN93olBb0ulMGgYI6/bk75u/MQGA5y538VF4eCs06Qjdr7PuWOhypVMVxQBfL2s7orf16H8XSacyuODZWZwgkPgXRlT221Kqwsp0r40x5gSwFGuuwSH70gL218N2t32A443GkcABuz3SRbtSSlWr5LRMOjy5gPs+z58nsOVAMmOmWoOik6/tyb+GdQBg/d4kTpy2Jh9e39d5pNNpO/0kzH8I3hkAiX+Blw90vapQueXMArUUlu1I5LXFOzhMCBno3Q2qZioxWRCRUHtEARHxB84HtgFzgZvsbjcBc+znc4ExIuIrIq2wJjKusi9ZpIhIP7FmA93ocIxSShWSlplNclom7/36N1GPzufqd5eTWdbyxy4MfnkpaZk5zF1/IG/Z6LjEk3n7h3VuSudm9QEYN2MNMc8tAqBvdCOn8+RNbPzrR3inP6x+D2JugYatCr3m4gfPBZzrLBw5mc7YD1Yx7Y94AKJD61X4vSlVFUpzGSIcmG7PW/AAZhljvhOR5cAsEbkN2AOMBjDGbBaRWcAWIAu4276MAXAn+bdOLrAfSinl0qgpf7A1ITlve1X8MXYfTaVNk8Byn/PBWeudah0cSUkn0NeLIynpeW0+Xh40CCj8V76vV4G/r3Ky4af/g+VvQWgHuHk+RLme8NiyYQDgPLKQVWCtiDeu0UmMqmYqMVkwxmwACv0PNsYcBYYWccwkYJKL9jVAcfMdlFK1WHpWNmmZOQT7V3w43RjjlCjk+nnb4XInC8t3HuXrddY86/M7NmHR1sN8sWYvtw5oxb+/2wLAskeGALh8D4PaNs57PrRDE8hKg01fW1UUh70APgFFvraXvT7EdxsSuHdoW4BC7y93NEOpmkYrOCqlKs0dH6/l5+2J7Hp+ePG1B0qw+UAS89YnOLV9fWd/rv7fCv7ce7xc50xMSefbP/cDMOas5lzRK5JFWw/zztKd7Eq0FnDq0bwBze0RAFfJQoCPFz9MGMSva2IZO6SDtQz0rT9ASFSp49h+yFpwauO+JG6Zttppn4eHLietaiZNFpRSlebn7YkA7D6aSlTj8l1/f2BWLLPX7c/bfv2aHvSNbkh4sD/tw4LIyCrfDVJnTbLmHbRpEsgLBVZy/GHzQbw9hW/uOjuvLcjPOVm4faA1D6FD2kY6bL0VDkXDLWVLFBx9sWZPuY5Tyh105RGlVKVIy8xfenl1/LFyneNwSppTogBwec8IwoP9AfD28iDptDXfYP3eE2TnlC5xOOywKmT7sCCXfRrW83EaDfH0EB6+qD3ensJl3ZvxaD9/mHUTTBsOXv4w8u0KLd70yQpNFlTtoSMLSqlK8cmK3XnPy7JYUq6cHEOfSYud2m7o53yr4qGkNA4mpzFp/hbeW7aLWwZE8chF1i2O/j6euFJwpCLu8EmX/bxc/OK/e0gb7h7SBr66DaZ8C54+cM4jMHCCdQmikozuHUmT+i5WmFSqhtBkQSlVKZ6zKxwCpGZk88Omg7z4wzbq+Xry7V0D8ib4OUpMSc+7PHDHua3z2vu2asjn4/oVuoZ/5GQ63mQxe8V2wIdf/0rks5V78PP2ZP1TFxY6/6s/bi80UvG/sb0hOxMSNrDuyQuY+uvfHP3tA6JOHoRFy+D0cTgaB8d3w83zrMsMbS+ABs2hz3io36wC3yXXXh7dvdLPqVRl0mRBKVVhjrUP/Lw9OJ2ZzR2frM1r23YwxWVZ5Km/7sx7/u4v1vMp1/dieNdw544ZqbB2Gj95vU4rD2txp63SghvTXiM9K4fX5FX46XdrNcam3fIKIU1eEgdAX9lKS4+DvDQ0BBZ9BruWQXoyDR+O46LOYZz4YyUDPDbBHx7gGwiN2kLUADix10oWuo+plO+TUrWVJgtKqQr78LddADx5SSfe/jmOHfaM/1xTlsYx5freedsnUjP4dOWeQnc8NAv2K5woZGXA230haQ+JtGd25iCy8cSDHBLt2gjeZFm1Dn5/HRq0AP8QDmf4AP8E4D/e79PaIwF+FatgUqdLof1w8AkkurEn3TP/RVh9X1Y+dn7lfmOKkFPKuRZK1RSaLCilKmzDPmu55uv7tuDD33bl3RWRa6m9feRkOsH+3vT49095+0ICvDmeapVTPpBkT0TMzoSt86DjpVbZ5D63Q0RvbvrgJKezsyloXOZDxE/sD1vnQtwiTHYGa7ceAwwPXdieXR5TyGzSmA7tO4Kn810Owd7w3o0xhAZV/ZyBqEYBxB9NJd2hMNMtA6Kq/HWVqihNFpRS5WaMIXbvCdbvO8HANo3x8/Zk/4nThfqlZmSTkZVDzHOLuDom0mlfbqIAMKxTY9gwC359GY78BaOmQvdrYMD9APz2r3QGvfQzqRmFEwYCGlqrOfa+mR2HUrhzw68AxEQ1pF9022LfxwWdwsr4zssn/mgqAPtPWF+fvrQTNw8oXBpaqZpGb51USpXbV2v3MWrKH+w7fhpvT9cFhQa2aUyQnxcJSVYSMWvNvrx2gJeu7Ma8O3ozuctO3jn1T5g9DsQTrvkUuo52OlejQF/XiQLW4lAAp9KzmLfeWqPuwQva0bdVw4q/0Up22L58UtQdHKr6iEhTEZkpIjtFZIuIfC8i7URkk7tjq0l0ZEEpVS6HktN4+KsNedv3DXX+67178wZ0bBqEt6cHv8Ud4dyXlzrt/8e50bx2TQ8aB/ogf/9M17gnrcmEV34AnUeBR/G/SIN8vUixF4EC2Hogmb7RjXji2018Y1dqHNEtvEKVJKtKarqV8Ph5a7LgTvaiht8A040xY+y2HkD1DDXVIjqyoJQql77/ya+JEP/CCHq2CHHaf0nXcF64shvr9hQuz+xFFoNO/0zoqpesX+atBsNN38G966xlnYtJFF67xrrNMLrA+hDX2MtLbz+YP7myYBXGmmLm6r0A+Guy4G5DgExjzLu5DcaYWGBv7raI+InIRyKyUUT+FJEhdntnEVklIrEiskFE2trtNzi0/89ehLHW05EFpVSZpaTlzzNY+tBgp319ohqyKv5Y3nLLXg61EppwnNGev3C91yKYfQwaRkOfcRDUFFoNKtVrj+oZiTHQq0UIYz9cyd5jznMkIkL82ZKQTK8WDWgc6FPOd1i1Fm21bv/UkQVL1KPzXwd6VPJpY+NfGDGhhD5dgLUl9LkbwBjTVUQ6AD+KSDvgDuANY8ynIuIDeIpIR+AaYIAxJlNEpgDXAzMq8kZqAk0WlFKllpqRxeuLdjD1178BePHKroXWgDhtl31uHGjdXZBpL8N8n+ds7vGegw+Z5EQNgrPvgzbnl6tk8hW9rEmSyx45D4CxH6wkOS2LlLRMftpyiCBfL2bfNaB8b7Ia6ZyFWmEg8CaAMWabiOwG2gHLgcdFJBKYbYzZISJDgd7Aavvylz9w2D1hVy5NFpRSpfL+sr+dqjQCjOwRUahfeLAfG/cnWas27v6D0Z2DeMZeill63gAD78WjUetCx1VEw3o+xB89RezeEwBOcxlqMr0MYSnFCEBV2QxcVUIfl5NejDGfichKYASwUERut/tON8ZMrNww3a/ElF5EmovIzyKyVUQ2i8j9dvvTIrLfvi4TKyLDHY6ZKCJxIrJdRC5yaO9tX/eJE5HJUhNnHimlnOTkGIa/saxQovDf0d1dDqO/PLo70y8LIWrxHfDRxdyc8zUrJg6lz80v4j3ydajkRAEgrL4fh5LT2X/cuiTx+bh+lf4aVUEvQ7jdEsBXRMblNojIWUBLhz6/Yl1KwL780ALYLiLRwN/GmMnAXKAbsBi4SkSa2P0biojjuWqt0owsZAEPGmPWiUgQsFZEciuqvGaMecWxs4h0AsYAnYFmwCIRaWeMyQbeAcYDK4DvgWHAgsp5K0qpqrBhfxJb7JGB6NB6fHPnAA6npNHW1eqN2ZkEL32Cc1e9B15+MHgicva9NPXxo2mwX5XFGFbfj4ysHJbtOEKAjydnRYWUfFAN4Oetc8zdyRhjRGQU8LqIPAqkAfHABIduU4B3RWQj1u/Dm40x6SJyDXCDiGQCB4F/G2OOicgTWPMaPIBMrDkPu6nlSkwWjDEJQIL9PEVEtgKFxx7zjQRmGmPSgV0iEgf0EZF4oL4xZjmAiMwALkeTBaVqrN1HT3HFlN8BWPLguUSHWncgBAe4uMsgMw2+uB7iFkGvm+C8JyCwSbXEGWav2Dh/YwKN6vm4XLSqJtLLEO5njDkAXO1iVxd7fxpws4vjngeed9H+BfBF5UbpfmX6iRKRKKAnsNJuuse+ZeRDEclN5SNwuO0E2Ge3RdjPC7YrpWqIrOwcUjOs6/2HktM49+Wl5Bi4sFNYXqJQiDGQeRq8fCFqEAx/BS6bXG2JAkCToPxRi6PlWB7bXfQyhKotSj3BUUQCga+BCcaYZBF5B3gWMPbX/wK34noyiCmm3dVrjce6XEGLFi1cdVFKVaJN+5O45M3f8rafvrQTT8/bAkDbJoE8MqyD6wP3rICFj0H0EBj6JAycUA3RFhbsXzPrKZRERxZUbVGqkQUR8cZKFD41xswGMMYcMsZkG2NygPeAPnb3fUBzh8MjgQN2e6SL9kKMMVONMTHGmJjQ0NCyvB+lVClk5xgOJadx9bvLWbYjkXs+W+e0PzdRAPjxn+fQpkABJPashM+ugQ8vguQD0KRjdYRdpCC/2nFj13OXd8l73j+6ER4eOsdb1Q4l/oTZdyx8AGw1xrzq0B5uz2cAGAXk1tGeC3wmIq9iTXBsC6wyxmSLSIqI9MO6jHEj9r2rSqnqde/n6/h+40EAxn6wCoDzOjTh4i5NnUo4f3xbH+dyyZlp1toNW+eCf0MY8jj0vxt8nGstVLf6DiMLt9bghZlCAvKLRD1/RVc3RqJU2ZQmHR8AjAU2ikis3fYYcK1dQ9tgzR79B4AxZrOIzAK2YM0cvdu+EwLgTmAaVqGKBejkRqWqjTGGjfuTWLLtcF6i4Ghsv5b0b92IL9fuY9WuYyx9aHB+waX0k1ZC4O0HfvVh8GNw9j1uTxJy1XMobvR/l3ZyYyTFcxxI8NRRBVWLlOZuiN9wPd/g+2KOmQRMctG+BnuGqVKqaizacojbZ6zh4YvaM+as5jQK9CU5LZNuT//o1O+z2/vSOyqEOz5eS6vGgQzpYE1InPWP/vmd0lNgxbvwx5tw+dvQ8VIY+XZ1vp1SqS0lWxzD9CpilU6laqLacaFPKVUq2TmG22esAeDlhdt5eeF27hrcmogQf6d+7cOCONteIvqjW/oUOg8Zp2D1+/D7ZEg9Au2GQUjNHd6vLWKi8pfL1pEFVZvUjpuRlVKl8tz8LYXapizdyePfWFOKpt9qJQaX9WhW9EkOb4M3e8NP/wdhneD2JXDdF9BUBwUrKne9DADPWjIaUteJSFMRmSkiO0Vki4h8LyLtRGRTyUdX6HUDRGS+iGyzqyO/4LDPV0S+sKsdr7TLFuTuu0lEdtiPm6oyRkc6sqBUHTJvvXWD0c7/DOfzVXt44lvnz7tz2jbms3F96R/dqPDB6SngG2SVY44aBDG3Qsv+hfvVUO9c36tW1S3wKscCWqpy2RP4v8Faz2GM3dYDCKumEF4xxvxsr1q5WEQuNsYsAG4Djhtj2ojIGOBF4BoRaQg8BcRgzRdcKyJzjTGF14GvZJosKFUHJKVmsmLXUY6czODx4R3x9BBu6NeSq3pHcuDEaZ79bgvPX9ENEeHs1o2dD045CMvfgjUfwW0/QlhnuPI997yRCri4a7i7QygTT52z4Ozp4KWu25MG2/tfx/Uy1hN4OimWp4NvxrHSYu5xxRsCZBpj3s1tMMbEFvhL3g9rqYIYrEn7D9i/4DsDHwE+WKP0V9orT94A3Ge3rwTucpjkn8cYkwr8bD/PEJF15JcXGAk8bT//CnjLTmwuAn4yxhyzY/sJa9mEz+0qyZ/Z78kbq1bR80Ab4GVjzLt2vaQ5QIjd5wljzJxSfJ80WVCqNjPGsP1QClN+3slce1Shb3T+dXE/b0+iQwNdz0tIPgC/vwGrP4CcLOh2jXU7pKoWehmiRugCrC2hz90AxpiuItIBa92HdsAdwBvGmE/tkQFPEekIXAMMMMZkisgUrEWoZhT3AiLSALgUeMNuyquEbIzJEpEkoBFFV0jOtdcY019EXsO683AA4Ie1uua7WGtfjLILKzYGVtgjEy4LJDrSZEGpWmjwyz8TfzS1UHuzYD+6RgSXfIK4RTDzeitJ6H4tDJgAjdtUfqCqSDrBsYCSRgKeTppQwv5pWL8gK9tA7JpAxphtIrIbaAcsBx4XkUhgtj2qMBToDay279DxBw4Xd3IR8QI+ByYbY/7ObXbRtTSVkOfaXzcCgcaYFCBFRNLshOQU8B8ROQfIwUo0wrAWwiqWJgtK1TJpmdkuE4URXcOZfG1P17cRZmVA7Kfg6Q09b4BGba2RhEEPQkidWEG31vHSZKEm2AxcVUIfl/9QxpjPRGQlMAJYKCK3232nG2MmliGGqcAOY8zrDm25lZD32clEMHDMbh/s0C8SWOqwnW5/zXF4nrvthTXKEQr0tkc+4rFGHkqkM2yUqmW22stFx7S01m6777w2xL8wgrev71X4r9W0ZPj1FXi1A3w3AbbOs9pDWlqLPWmiUO1G9bRGjbXUc42wBPAVkXG5DSJyFuD4g/Er1i9Z7MsPLYDtIhIN/G2MmYz1F303YDFwlYg0sfs3FJEif8hE5DmsRGBCgV1zgdw7Ha4CltiXChYCF4pIiL1444V2W2kFA4ftRGFIgfdZLB1ZUKqGMMaUWFzox80HGf+xdYn1oYva06N5A3y9XOT8xsCvL8OKKXD6OLQ5H/rdBa3Pq4rQVRm8fFU3nhnZ2d1hKMAYY0RkFPC6iDyKdU0/Hudf3lOAd0VkI9YEx5uNMekicg1wg4hkYg3j/9sYc0xEnsCa1+ABZGLNedhd8LXtyxePA9uAdfbP/lvGmPexllj4WETisEYUxtjxHhORZ4HV9mn+nTvZsZQ+BeaJyBog1n7tUpFSzGtwq5iYGLNmzRp3h6FUlYnde4JFWw7x1s9xeW23DmjFOe0aM6htaN5oQVpmNh2e/CGvz7ZnhxW+VTA7E8QTPDxgxkjw8odzH4aI3tXyXpQqBR1SqYU0WVDKjR74IpbZf+4vts9ZUSF8dEsf7vv8T5ZsO4ynh/DNXWfTLbJBfqfkA/Dnp/DnDLjkdWgzFDJPg7d/UadVyl00WaiF9DKEqvPSMrMRAV+vmlewxzFRuKJnBKNjmnPteyuc+qyOP06Xp/IvS2546kLq+XpZ8xG2zYcNX8CuX8DkWMWUchd30kRBqRrBngjpW6B5rDFmozviKQ9NFlStt37vCRZvPcRdQ9rkDcunZWbzzLwt7DueyrIdRwB46apuPPr1Bsb2a8k57UKZ8EUs02/tQ68WIcWePy0zu9IqA+bkGD74bRcD2jRm15FTADxzWWfG9muJiLUg0sA2jfkt7ojL42/o18JKFAA2fwPz7oP6kTDwAeg1FkKiKiVOpVTlMcb0dXcMFaWXIVStcjgljUb1fPOu4y/cfJB/fOxcU+WpSzvxzLzCayQ46hfdkBV/588LuqJXBK9e3aNQvzmx+7l/Zixz7h7A4q2HEBH+eUG7Usc7J3Y/s9bs5c1rexHg4+k05yDX6sfPJzQo/4+OIyfTmTAzlmdGdia6cT0+XWmVbfYii3e77eT8sFMw9EnIzoKEWGs+ghb4UbWH/methTRZUDXenNj9LNtxhB7NG/DEt5v4z6iuXNe3BQCDXlrC3mOny33uphzlPM9YWsghbh1+Dj7hnaBpV/BvQHpWNu2fKPzL/d0berPn2CmuiWlBcIC3y/OmZWazJv44N3ywMq+tcaAvR06mF+q76/nhRd8FYQwc+5vfF3xKyx3TiZQj0KI/3Py9NYlRqdpHk4VaSJMFVaM9PXcz0/6IL9TeLTKYDfuSALhzcGveWbqzUJ+bz45i5uo9fH/fIEKDfLn41SU0OrWDrBxhs4lisEcsH3i/jKcYMo0n3mKXb+80Eq6ewfe/ryXwh/sJlROEyEkCSOckfmzOiWJc5kNc2r0Zb/ZNhvDu4N8g73VPpmfRZ9IiUjMKlYMHrAmLq+OtdV9GdAvn7et6WTtSj8HBDXBsF8TcYrVN6Q+HrVGShOAeNLrgIXw6jdBEQdVmmizUQiXOWRCR5lh1rZtiVYGaaox5w1796gsgCuu+1KtzV74SkYlYq2ZlA/cZYxba7b2xynH6A98D95emJrU6My3bkegyUQDyEgWApvX98PYUMrPz/ytte3YYfl4ePN3HwNZ3YcdP/Jb9J3in83nWECZmjWPgkOEcSEljZ9gwbp53nMW3t6M1e8C3PgCLNu9jrJzmqHc4G9IDSMWXQE5zgkAA/LJPwaejAQPthlllk1sN4tGvdzglCue2C+WXvxLztj8f1482jy8ADFeGH4WfnoLt38ORv6wOHl7Q60bw8ITet1hVF1v0J7xJh8r5xiqlVBmVOLIgIuFAuDFmnYgEYS26cTnW6l7HjDEv2MUsQowx/xKRTlh1rvsAzYBFQDtjTLaIrALuB1ZgJQuT7eU4i6QjC3VfWmY2KWlZJKak06mZ9Yt6Tfwxrnp3OQCvjO7OQ1+uL/L4Lf++iKMnM3hxwWZeHZCFj1+gdSlh+wL4fIzVqVlPaDkAmvUkPaIvi/Z5cXGXpnh4CL/8lchNH67i6zv707ultZDSuj3HuWLKHwCc3zGMRVsP8dEtZ3HLR6sdXtkQf084bPoa/vwE0pPI8fBhUWZXxmc+CMAzAbMYe04ndp304o0/jnJZr5ac3zmCUT+HELvnGNsaPIBv+lFodY71aNbTWvUxsEklf5eVqjF0ZKEWKnFkwRiTACTYz1NEZCvW4hMjya9RPR2rPvW/7PaZxph0YJddgaqPXYO6vjFmOYCIzMBKOopNFlTdkZNjEIFTGdkIUM/Xi6H/XcrOxFN5fd67MYYLOoUxfblV8GzK9b0Y3jU8L1mo5+PJKafhfUPAoXUErH6ft/YuhGknoMf1cPkUaHk2XPaWVb2wfv7yxb7ACIfFFQN9rTsd/k48xan0bAJ8PPMSFYBJo7rQL7ohg9uFEuDj6TBqIKSF9cAvsjcM/T9y9q5i6of/I9X48fjwjowb2BJeuguWfktrYLIPsAnY7EHb6Nn8iQcbB0wmplcfqNeosr7NSilV6cp066S9xndPrDW6w+xEAmNMQm4tbKxEwvFG8dwlNDPt5wXbXb3OeKy1uGnRokVZQlQ12L+/2+J0WeGT2/o6JQoA42asYeVjQ5lnL7c8vKv1S/6y7s2Yu/4Adw1pw5Sf4ziVkU172cPskDfhg/3WpYOOl0HrIRA92DqZX7B1O2EJAn2tSYoPf7Wh0L659wwgrL4ftw+KBmD9UxeyYNNBDiWlMen7raSkZVm3VXr7sbNeT17Iug6Arf1aWpcRHt1tVVVMS7LKLmdngpcvD3s3IyD4b7r27wA1sP6DUko5KnWyICKBwNfABHst7CK7umgrzdKa+Y3GTMVaiYuYmBid01CLGWNYuj2RW6atLrTvVrtt1WND+WTFbiYvscod9/3PYgB8PPMn8d3aL5zkjd9zfcI33N2/EVz8ImScgq9/hfYTofPl4BtUrhgD/Vz/GNw2sJVzlUTA29ODy7o3Y06sVUwpOS0z77bHxduslWjP69AEfx+HBMDTG+o1th62UODpy3R9AKVU7VCqZEFEvLEShU+NMbPt5kMiEm6PKoSTv2Z37tKauSKBA3Z7pIt2VYet2X28UKLQPTKY9fuSyMjOwdtTaFLfjwcubI+nhwevLforr9+jF3eAw9tg9fv0WP8503xOwp5gaHCN1cGnHlz7eYVjDPR1/WNwVe9Il+0AQXaC8crC7ew5lsrtg1rxwgJrTZaHL2pf4ZiUUqomKc3dEIK1AtZWY8yrDrtyl9B8wf46x6H9MxF5FWuCY1tglT3BMUVE+mFdxrgReLPS3omqcbJzDKMdrv0ve2QIzRsGkJ1jaP3Y9wD8a1j+DP/7z2/LW4u20FwO87dpxuBW/jC1P2Cg46XQdbR1icGrYNXUiikqWXAslFRQsL8PAAs2HQTgn1/kT8CMDNEyy0qpuqU0IwsDgLHARhGJtdsew0oSZonIbcAeYDSAMWaziMwCtmAt53m3MSZ3Rtid5N86uQCd3FhnZGXnMHvdfi7pHo63pwc/bj7E3Z+tA+De89pw73lt8bGXUvb0EEZ0DWf+xgTG9GkBmWmwYyFs/IrN9RZhsjLpmP4R0RFNYfRHEHmW0xB+ZcutBgnw1nU9ueezPwEICfAp8pgezRu4bP/u3oEE+bku1KSUUrVVae6G+I2ib3UZWsQxk4BJLtrXAF3KEqCq2d5f9jfPzd+at/3I14UnCT5wQbtCFQrfuq4nr+f0wHvJM7D6A8hIgcCmJDS/hFf+aoIHOVbH9hdXafwFXdKtGc/M20JiSrpTElGQq30dmgbRJSK4KsNTSim30IWkVKlstOcY7Dpyigs6hhEc4M3P2w47JQquDGrb2DlRMAb+Woi0OR9vTy9ISYAOI6DbaIgegldyBvNeWFLkX+5VqVuk9Yv+k9v6ciI1o8T+H94cw63T8muANCii9LNSStV2Wu5ZlWhrQjIXv7EsbzuqUQCjY5rz8sLtLvvHtAzhnRt60zjQJz9RyMmBLd/C8rdh/xoYPQ06j7KShwKjDr/HHaFtk0Ca1PerondUWFpmNp4egrdn6csoG2P4ZOUemgX7cdv0NUSG+PPbv86rwiiVqhO0KFMtpCMLyqWMrBy+WLOX7Owcni6wgmNCUlpeotC3VUPuGNyarhHBPDNvC/PWH2BMnxbOkwPjf4MfHoWDGyGkFVz6BnS4xNrn4hbcAW2qbn5CUcqzBLWIMLZfS9IyrSk5dw1uU9lhKaVUjaDJgiokLTO70FLKPVs04MUru3H9+ytJTLFWThzSPpSPbumT1+eRi9rTrIEfI3s0yx8x+PsXmHEZ1I+AK96DLldaxYrqED9vT+JfGOHuMJRSqsro0nXKyY+bDxZKFAC6NAumXVgQ/aLzyxLfOrCVU5/mDXyZGB2P9+xbYdaNVmPUIBg5Be5ZA92urnOJglJKnQl0ZEHl+T3uCOM/Xpu3/dntfckxMGVpnFUgCXjt6u55pZhbh1qrL5J5GtbPhBVTrJUTAxpBt2us0QUPD+h5fbW/F6WUUpVHk4Uz3InUDNbEH+f8TmFMcrizYWiHJpxtzx0Y2DZ/DoGXp0feYkph9f2siYtT+sHxeGjaDa760FqjwVPvDFBKqbpCk4UzzNdr95Gamc3Yfi0BeGPxDj76PZ4v7+jPloRkAnw82fLvYUWf4PRx/rxwJx6bZ+OZ1h8CGsJ5T0JQU2sJ6KLXDFFKKVVLabJwhth7LJVBL/2ctx275wT/vbo7y3ceBcgry5xT1K20KYfgj8mw+n18s9IgvDskH7CSha5XVXn8Siml3EeThTrosW82kp1teOqyTgT4WP/Ew17/1anP1+v2MbJHM7YdTHFq/+qOswufcO10mP8gmGxrfYb+d1vJglJKqTOC3g1Rx8QdPslnK/fwxZq9dPq/hcQ8t4jFWw9xKiOb0CBfdj0/PO8SxI0frgLgj0fPy7t60DbMnrSYuB32rLSeR/SGnjfA3avhiqmaKCil1BlGKzjWMVGPzi9y3xtjejCyRwQb9p3gsrd+B6B1aD0WPziY/SdOE3/kFAMan4alz0PsZxA1EG7+rrpCV0qdGXRiUy2kIwt1yNaE5Lznl3VvVmj/xV3CAegW2YB59wwk2N+bz8f1AyDCO5UBf78Ob/aGjV/B2ffAVR9VS9xKKaVqNp2zUAds2p/EyLd/JzvHGiX67t6BtA4NJDElneV/WxMYb+jXIm+JaICukcGsf+pCa8MYmDYCErdBj+tg8ERo0Lza34dSSqmaSZOFWmzqrzv5z/fbCrW3aRKIn7cnn4/vl3dZ4p4hbZ07Je2Hle9Cn3HQoAVc/CIEhkGTDtURulJKqVqkxMsQIvKhiBwWkU0ObU+LyH4RibUfwx32TRSROBHZLiIXObT3FpGN9r7JInpDfkXMXrfPZaLwy8ODnRZFWv/Uhbx7Q2+aBtsrOB6Phzl3wxvdrBUgd/xktUefq4mCUkopl0ozsjANeAuYUaD9NWPMK44NItIJGAN0BpoBi0SknTEmG3gHGA+sAL4HhgELKhT9GSYrO4fZf+7nka82OLWf37EJi7YeJjTIl5aN6jntC/b3ZliXptZIwsKJsGWuVV0x5jbofxeERFXjO1BKKVUblZgsGGN+FZGoUp5vJDDTGJMO7BKROKCPiMQD9Y0xywFEZAZwOZoslNqh5DT6/mexU9sDF7TjvqFtXR9gDOxfC6lHoZ09wBO3BAY9YCUKwRFVHLFSSqm6oiJzFu4RkRuBNcCDxpjjQATWyEGufXZbpv28YLtLIjIeaxSCFi1aVCDEuiEjK4fr3ltRqH1UTxffwpxs2PItLHsNDm2Exu2g7YVWcvDwDvD2r/qAlVJK1SnlTRbeAZ4FjP31v8CtuL5/1hTT7pIxZiowFaw6C+WMsUab/kc86/Ycp1eLEA4mp9EpvD6Xurjd8fuNCdz16ToAHhnWnst7RJCWmc0vfyXSvGFAfsecHFg11Zq0eHwXNOkEw1+xVn/MnR6iiYJSSqlyKFeyYIw5lPtcRN4Dciv37AMc77mLBA7Y7ZEu2s9IKWmZPDV3MwBzYvO/DSO6huPhkZ9XORZY6tmiAeMHRePlac1Jjc5dHjr5AASFW0tBb/kWApvA0P+DTiPBI3+io1JKKVVe5UoWRCTcGJNgb44Ccu+UmAt8JiKvYk1wbAusMsZki0iKiPQDVgI3Am9WLHT32n4whWl/xHPveW1o1qB0f7GnpGXy4W/x/HUoxeX+6Me+p3fLEP47ujsPfrnead/Xd5ydn0jkZMOeFbDyHdg2H25dCM37wPVfgm9Qhd6XUkopVVCJyYKIfA4MBhqLyD7gKWCwiPTAupQQD/wDwBizWURmAVuALOBu+04IgDux7qzwx5rYWGsnN94+fQ2LtlqDK/tPnKaejycLNh1kzRPn0zjQ1+Uxx09lcPEbyziYnJbXlrsmw4zlu3ln6U4A1u4+zuBXlgKGUJKIlER2mAgrUdj4FaybDvvWQGYq+IfAgAlWnQTQREEppVSVKM3dENe6aP6gmP6TgEku2tcAXcoUXQ30xeo9eYkCwK9/JeY9/z3uCCN75E86TErNZNfRU6zadbRQTYTLujfLG5H417AOtGpcj53fTGKAxyYi5AiRHkfwJROALefbd62mHoO0ZOg51lrcqeOl4BOAUkopVZV0IalS+uWvRG6yV2kEuLxHMzJzDPM3JOS1RTeux5KHBgOwcPNB/vHx2kLniX9hhPUkIxW2zIF1M2DMpxDQkA3v3oI5EMteE8olg/pAg5YQ3Ny6xBDQsErfn1JKVRMtyFcLabnnUvhyzV4eLlAI6clLOnHidCbzNyRwU/+W7D9xmkVbD7NsRyIN6/m4TBReu6qTVTFx02zY9h2kJ0PDaDh5CAIa0mX8h0Q/9j1Xx0RyyYW6DLRSSqmaQUcWSvDUnE1MX74bgA5Ng3hldHc27Eviur7O9R8Skk7T//klhY73IAdPcujbpikzmn2Fx6qp4BsMHS+B7tday0A7VL4+nZGNj5cHnh6afCul6iT9cKuF6vTIgjGGaX/Ec0XPSIIDvMt8/IZ9J/IShVE9I3jtmh4AdIkILtQ3PDj/johQTnCOxwbO9VzPxQHb8B42CXr2hYP1oPV51sPL9URIfx+93VEppVTNUqeThbjDJ3lm3ha2JiTz0lWuh/Uf/XoDM1fv5e3rejGiW3he+1+HUrjsrd8BeP/GGAa3Dy3x9V5o9xcDd79FpBwBwASGIdEXQaPWVoemXayHUkopVYvU6WShbVgQvVuGsOdYapF9Zq7eC8Ddn63j2KnOPDlnM54eQnaOdXnm4Yvac36nsFK93phzunF02VnMOBbJBcOvIrxDX6dLDEoppVRtVKeTBYCQAG8OnEhzuW9rQrLT9pNzrKqKuYkCwN1D2pT+xdoMpVGbodxY9jCVUkqpGqvOJwv1/b3ZmpDCqfQsHvl6A10jgjmdkc2E89ty8RvLij02btLF1RSlUkopVXPV+WQh2N+b/SdO0/mphQB5dREcF2Ea0TWc05nZLNl22OnY3HUYlFJKqTNZnU8WQgJ8XLY/ZK+9cFP/ljwzsgtzYvezZNthgvy8+P6+Qfh6aaKglFJKwRmQLJT0S3/i8I4ARNill89pF+q89LNSSil1hqvzyUI9X+e3GFbfl0PJ6Xnbft5WXYOeLUJ49OIOXNq9WbXGp5RSStV0dX6sPcJh+eiBbRqzYuLQvG1vz/zbGj09hDvObe3UXymllFJnwMjC4PahXNU7knPahTKoTWNEhG3PDqPDkz/QPEQvNyillFIlOWPXhvh5+2HahwXlLROtlFKqWmilulqoxMsQIvKhiBwWkU0ObQ1F5CcR2WF/DXHYN1FE4kRku4hc5NDeW0Q22vsmi7i3tOGQ9k00UVBKKaVKoTRzFqYBwwq0PQosNsa0BRbb24hIJ2AM0Nk+ZoqI5K6M9A4wHmhrPwqeUymllFI1UInJgjHmV+BYgeaRwHT7+XTgcof2mcaYdGPMLiAO6CMi4UB9Y8xyY133mOFwjFJKKaVqsPLeDRFmjEkAsL82sdsjgL0O/fbZbRH284LtSimllKrhKvvWSVfzEEwx7a5PIjJeRNaIyJrExMRKC04ppZRSZVfeZOGQfWkB+2vuogr7gOYO/SKBA3Z7pIt2l4wxU40xMcaYmNDQ0HKGqJRSSqnKUN5kYS5wk/38JmCOQ/sYEfEVkVZYExlX2ZcqUkSkn30XxI0OxyillFKqBiuxzoKIfA4MBhoDh4CngG+BWUALYA8w2hhzzO7/OHArkAVMMMYssNtjsO6s8AcWAPeaUhR5EJFEYHeZ31nlaAwccdNrO9I4nGkczjQOZxqHs5oWxxFjjN4NV8vU+KJM7iQia4wxMRqHxqFxaBwaR92IQ5VPnV8bQimllFIVo8mCUkoppYqlyULxpro7AJvG4UzjcKZxONM4nGkcqsJ0zoJSSimliqUjC0oppZQqliYLSimllCrWGZUsiEhzEflZRLaKyGYRud9ud7nktog0svufFJG3HM4TICLzRWSbfZ4X3BGHve8HEVlvn+ddh1U+qzUOh3POdVzO3A3fj6X28uix9qOJq9eshjh8RGSqiPxl/z+5srrjEJEgh+9DrIgcEZHX3fT9uFasJeo32P9nG7spjmvsGDaLyEuljaGccVwgImvt971WRM5zOFdvuz1ORCaLiKuS+NURxyQR2SsiJ8vyvajMOKSCn6eqmhhjzpgHEA70sp8HAX8BnYCXgEft9keBF+3n9YCBwB3AWw7nCQCG2M99gGXAxdUdh72vvv1VgK+BMe6Iw95/BfAZsMkd/y72vqVAjDv/f9j7ngGes597AI3d9e/icN61wDlu+HnxwioJ39jefgl42g1xNMIqIhdqb08HhlZhHD2BZvbzLsB+h3OtAvpj/dwuoGo/P4qLo599vpPV8PPiMg4q+Hmqj+p5uD0At755q+T0BcB2INxuCwe2F+h3M8V/CL8BjHNnHIA3MA+4xh1xAIHAb/aHRZmShUqOYynlTBYqOY69QD13x+Gwr60dk1R3HPb/zUSgJdYvx3eB8W6I4yxgkcP2WGBKVcdhtwtwFPC1+2xz2Hct8L/qjqNAe5mThaqIw95Xoc9TfVTN44y6DOFIRKKwMt2VFL3kdmnO0wC4FFjsrjhEZCHWX24pwFduiuNZ4L9AanlevxLjAPjIHnZ/sizDu5UVh/1/AuBZEVknIl+KSFh1x1HAtcAXxv40rs44jDGZwJ3ARqwF5DoBH1R3HEAc0EFEokTEC7gc54XvqjKOK4E/jTHpQATW4nq59tlt1R1HpamsOCr6eaqqzhmZLIhIINaQ/QRjTHIFzuMFfA5MNsb87a44jDEXYWXwvsB5JXSv9DhEpAfQxhjzTVmPrcw4bNcbY7oCg+zHWDfE4YW1survxphewHLgFTfE4WgM1v/VMquE/x/eWMlCT6AZsAGYWN1xGGOO23F8gTXUHY+1hk2VxiEinYEXgX/kNrkKzw1xVIrKiqOin6eqap1xyYL9wfU18KkxZrbdXNSS2yWZCuwwxrzu5jgwxqRhrfo50g1x9Ad6i0g81qWIdiKy1A1xYIzZb39NwZo/0ccNcRzFGmHJTZ6+BHq5IY7cc3UHvIwxa8sSQyXG0QPAGLPTHtmYBZzthjgwxswzxvQ1xvTHGi7fUZVxiEgk1v+DG40xO+3mfVjJZK5IrBGX6o6jwio5jnJ/nqqqd0YlC/aQ9AfAVmPMqw67ilpyu7hzPQcEAxPcFYeIBDr8UHoBw4Ft1R2HMeYdY0wzY0wU1sSyv4wxg6s7DhHxEnuWvf0hdglQ6jszKvH7YbDmjwy2m4YCW6o7DgfXUo5RhUqMYz/QSURC7e0LgK1uiAOx746xZ+jfBbxfVXHYQ+rzgYnGmN9zO9tD8yki0s8+542lib2y46ioyoyjIp+nqpq4e9JEdT6wfpEZrGHQWPsxHGuW9GKsvzIWAw0djokHjgEnsf4i6IT1l4DB+sDLPc/tbogjDFhtn2cz8CbWX5DVGkeBc0ZR9rshKuv7UQ9rxn/u9+MNwNMd3w+syXy/2udaDLRw178L8DfQwV0/L3b7HVg/LxuwEqlGborjc6zEbQtluHOoPHEATwCnHPrGAk3sfTFYiexO4C3KMPG0kuN4yf7+5Nhfn67uOKjg56k+queh5Z6VUkopVawz6jKEUkoppcpOkwWllFJKFUuTBaWUUkoVS5MFpZRSShVLkwWllFJKFUuTBaWqiIhk22WnN4u1MugDIlLsz5xdivi66opRKaVKQ5MFparOaWNMD2NMZ6xCRMOBp0o4JgrQZEEpVaNonQWlqoiInDTGBDpsR2MV0WqMVbTpY6xCUgD3GGP+EJEVQEdgF9YSypOBF7CqQfoCbxtj/ldtb0IppdBkQakqUzBZsNuOAx2wVgjNMcakiUhb4HNjTIyIDAYeMsZcYvcfj1Vt7zkR8QV+B0YbY3ZV53tRSp3ZvNwdgFJnmNwVB72Bt+wVO7OBdkX0vxDoJiJX2dvBQFuskQellKoWmiwoVU3syxDZWKvwPQUcArpjzR1KK+ow4F5jzMJqCVIppVzQCY5KVQN7xcV3gbeMde0vGEgwxuQAYwFPu2sKEORw6ELgTnsVTUSknYjUQymlqpGOLChVdfxFJBbrkkMW1oTG3KV8pwBfi8ho4Ges1fjAWsEvS0TWA9OwVs6MAtbZSwInApdXT/hKKWXRCY5KKaWUKpZehlBKKaVUsTRZUEoppVSxNFlQSimlVLE0WVBKKaVUsTRZUEoppVSxNFlQSimlVLE0WVBKKaVUsf4f3fewSkkycQ0AAAAASUVORK5CYII=",
      "text/plain": [
       "<Figure size 535x216 with 1 Axes>"
      ]
     },
     "metadata": {
      "needs_background": "light"
     },
     "output_type": "display_data"
    }
   ],
   "source": [
    "import matplotlib.pyplot as plt\n",
    "import seaborn as sns\n",
    "sns.relplot(data=sp500_history[['Close', 'Close_200ma']], kind='line', height=3, aspect=2.0)\n",
    "plt.savefig('chart.png')"
   ]
  },
  {
   "cell_type": "markdown",
   "metadata": {},
   "source": [
    "# Excel"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 3,
   "metadata": {},
   "outputs": [],
   "source": [
    "# Import libraries\n",
    "from openpyxl.chart import LineChart, Reference\n",
    "\n",
    "# 1. Set up an ExcelWriter\n",
    "with pd.ExcelWriter('excel_report.xlsx', engine='openpyxl') as writer:\n",
    "# 2. Export data\n",
    "    sp500_history.to_excel(writer, sheet_name='historical_data')\n",
    "    sp500_history_summary.to_excel(writer, sheet_name='historical_data_summary')\n",
    "    \n",
    "# 3. Add a line chart\n",
    "    # Point to the sheet 'historical_data', where the chart will be added\n",
    "    wb = writer.book \n",
    "    ws = wb['historical_data'] \n",
    "    # Grab the maximum row number in the sheet\n",
    "    max_row = ws.max_row\n",
    "    # Refer to the data of close and close_200ma by the range of rows and cols on the sheet\n",
    "    values_close = Reference(ws, min_col=5, min_row=1, max_col=5, max_row=max_row)\n",
    "    values_close_ma = Reference(ws, min_col=7, min_row=1, max_col=7, max_row=max_row)\n",
    "    # Refer to the date\n",
    "    dates = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=max_row)\n",
    "    # Create a LineChart\n",
    "    chart = LineChart()\n",
    "    # Add data of close and close_ma to the chart\n",
    "    chart.add_data(values_close, titles_from_data=True)\n",
    "    chart.add_data(values_close_ma, titles_from_data=True)\n",
    "    # Set the dates as the x axis and format it\n",
    "    chart.set_categories(dates)\n",
    "    chart.x_axis.number_format = 'mmm-yy'\n",
    "    chart.x_axis.majorTimeUnit = 'days'\n",
    "    chart.x_axis.title = 'Date'\n",
    "    # Add title to the chart\n",
    "    chart.title = 'Close prices of S&P 500'\n",
    "    # Refer to close_ma data, which is with index 1 within the chart, and style it\n",
    "    s1 = chart.series[1]\n",
    "    s1.graphicalProperties.line.dashStyle = 'sysDot'\n",
    "    # Add the chart to the cell of G12 on the sheet ws\n",
    "    ws.add_chart(chart, 'G12')"
   ]
  },
  {
   "cell_type": "markdown",
   "metadata": {},
   "source": [
    "# HTML"
   ]
  },
  {
   "cell_type": "markdown",
   "metadata": {},
   "source": [
    "### HTML"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 4,
   "metadata": {},
   "outputs": [],
   "source": [
    "# 1. Set up multiple variables to store the titles, text within the report\n",
    "page_title_text='My report'\n",
    "title_text = 'Daily S&P 500 prices report'\n",
    "text = 'Hello, welcome to your report!'\n",
    "prices_text = 'Historical prices of S&P 500'\n",
    "stats_text = 'Historical prices summary statistics'\n",
    "\n",
    "\n",
    "# 2. Combine them together using a long f-string\n",
    "html = f'''\n",
    "    <html>\n",
    "        <head>\n",
    "            <title>{page_title_text}</title>\n",
    "        </head>\n",
    "        <body>\n",
    "            <h1>{title_text}</h1>\n",
    "            <p>{text}</p>\n",
    "            <img src='chart.png' width=\"700\">\n",
    "            <h2>{prices_text}</h2>\n",
    "            {sp500_history.tail(3).to_html()}\n",
    "            <h2>{stats_text}</h2>\n",
    "            {sp500_history_summary.to_html()}\n",
    "        </body>\n",
    "    </html>\n",
    "    '''\n",
    "# 3. Write the html string as an HTML file\n",
    "with open('html_report.html', 'w') as f:\n",
    "    f.write(html)"
   ]
  },
  {
   "cell_type": "markdown",
   "metadata": {},
   "source": [
    "### HTML with template"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 5,
   "metadata": {},
   "outputs": [],
   "source": [
    "# 1. Create a template file as report_template.html\n",
    "\n",
    "from jinja2 import Environment, FileSystemLoader\n",
    "\n",
    "# 2. Create a template Environment\n",
    "env = Environment(loader=FileSystemLoader('templates'))\n",
    "\n",
    "# 3. Load the template from the Environment\n",
    "template = env.get_template('report_template.html')\n",
    "\n",
    "# 4. Render the template with variables\n",
    "html = template.render(page_title_text='My report',\n",
    "                       title_text='Daily S&P 500 prices report',\n",
    "                       text ='Hello, welcome to your report!',\n",
    "                       prices_text='Historical prices of S&P 500',\n",
    "                       stats_text='Historical prices summary statistics',\n",
    "                       sp500_history=sp500_history,\n",
    "                       sp500_history_summary=sp500_history_summary)\n",
    "\n",
    "# 5. Write the template to an HTML file\n",
    "with open('html_report_jinja.html', 'w') as f:\n",
    "    f.write(html)"
   ]
  },
  {
   "cell_type": "markdown",
   "metadata": {},
   "source": [
    "### HTML to PDF"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 6,
   "metadata": {},
   "outputs": [],
   "source": [
    "from weasyprint import HTML, CSS\n",
    "css = CSS(string='''\n",
    "    @page {size: A4; margin: 1cm;} \n",
    "    th, td {border: 1px solid black;}\n",
    "    ''')\n",
    "HTML('html_report_jinja.html').write_pdf('weasyprint_pdf_report.pdf', stylesheets=[css])"
   ]
  },
  {
   "cell_type": "markdown",
   "metadata": {},
   "source": [
    "# PDF"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 7,
   "metadata": {},
   "outputs": [],
   "source": [
    "def output_df_to_pdf(pdf, df):\n",
    "    # A cell is a rectangular area, possibly framed, which contains some text\n",
    "    # Set the width and height of cell\n",
    "    table_cell_width = 25\n",
    "    table_cell_height = 6\n",
    "    # Select a font as Arial, bold, 8\n",
    "    pdf.set_font('Arial', 'B', 8)\n",
    "    \n",
    "    # Loop over to print column names\n",
    "    cols = df.columns\n",
    "    for col in cols:\n",
    "        pdf.cell(table_cell_width, table_cell_height, col, align='C', border=1)\n",
    "    # Line break\n",
    "    pdf.ln(table_cell_height)\n",
    "    # Select a font as Arial, regular, 10\n",
    "    pdf.set_font('Arial', '', 10)\n",
    "    # Loop over to print each data in the table\n",
    "    for row in df.itertuples():\n",
    "        for col in cols:\n",
    "            value = str(getattr(row, col))\n",
    "            pdf.cell(table_cell_width, table_cell_height, value, align='C', border=1)\n",
    "        pdf.ln(table_cell_height)"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 8,
   "metadata": {},
   "outputs": [
    {
     "data": {
      "text/plain": [
       "''"
      ]
     },
     "execution_count": 8,
     "metadata": {},
     "output_type": "execute_result"
    }
   ],
   "source": [
    "from fpdf import FPDF\n",
    "\n",
    "# 1. Set up the PDF doc basics\n",
    "pdf = FPDF()\n",
    "pdf.add_page()\n",
    "pdf.set_font('Arial', 'B', 16)\n",
    "\n",
    "# 2. Layout the PDF doc contents\n",
    "## Title\n",
    "pdf.cell(40, 10, 'Daily S&P 500 prices report')\n",
    "## Line breaks\n",
    "pdf.ln(20)\n",
    "## Image\n",
    "pdf.image('chart.png')\n",
    "## Line breaks\n",
    "pdf.ln(20)\n",
    "## Show table of historical data\n",
    "### Transform the DataFrame to include index of Date\n",
    "sp500_history_pdf = sp500_history.reset_index()\n",
    "### Transform the Date column as str dtype\n",
    "sp500_history_pdf['Date'] = sp500_history_pdf['Date'].astype(str)\n",
    "### Round the numeric columns to 2 decimals\n",
    "numeric_cols = sp500_history_pdf.select_dtypes(include='number').columns\n",
    "sp500_history_pdf[numeric_cols] = sp500_history_pdf[numeric_cols].round(2)\n",
    "### Use the function defined earlier to print the DataFrame as a table on the PDF \n",
    "output_df_to_pdf(pdf, sp500_history_pdf.tail(3))\n",
    "## Line breaks\n",
    "pdf.ln(20)\n",
    "## Show table of historical summary data\n",
    "sp500_history_summary_pdf = sp500_history_summary.reset_index()\n",
    "numeric_cols = sp500_history_summary_pdf.select_dtypes(include='number').columns\n",
    "sp500_history_summary_pdf[numeric_cols] = sp500_history_summary_pdf[numeric_cols].round(2)\n",
    "\n",
    "output_df_to_pdf(pdf, sp500_history_summary_pdf)\n",
    "# 3. Output the PDF file\n",
    "pdf.output('fpdf_pdf_report.pdf', 'F')"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": null,
   "metadata": {},
   "outputs": [],
   "source": []
  }
 ],
 "metadata": {
  "kernelspec": {
   "display_name": "Python 3",
   "language": "python",
   "name": "python3"
  },
  "language_info": {
   "codemirror_mode": {
    "name": "ipython",
    "version": 3
   },
   "file_extension": ".py",
   "mimetype": "text/x-python",
   "name": "python",
   "nbconvert_exporter": "python",
   "pygments_lexer": "ipython3",
   "version": "3.7.6"
  }
 },
 "nbformat": 4,
 "nbformat_minor": 4
}
